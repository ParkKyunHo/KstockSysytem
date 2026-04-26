# -*- coding: utf-8 -*-
"""
V7 Purple 3분봉 백테스트 - Excel 출력기

openpyxl을 사용한 전문적인 백테스트 보고서 Excel 출력
"""

from datetime import datetime
from typing import List, Dict, Any, Optional
from pathlib import Path
import pandas as pd

from .config import BacktestConfig, Trade

# openpyxl 스타일 (모듈 레벨 import)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelExporter:
    """Excel 보고서 출력기"""

    def __init__(self, config: BacktestConfig, logger):
        self.config = config
        self.logger = logger

    def export_full_report(
        self,
        trades: List[Trade],
        report: Dict[str, Any],
        filename: Optional[str] = None
    ) -> Path:
        """
        전체 백테스트 보고서 Excel 출력

        Args:
            trades: Trade 리스트
            report: 분석 보고서 딕셔너리
            filename: 파일명 (기본: v7_backtest_report_YYYYMMDD.xlsx)

        Returns:
            저장된 파일 경로
        """
        if not OPENPYXL_AVAILABLE:
            self.logger.warning("openpyxl 미설치. CSV로 대체 저장")
            return self._export_csv_fallback(trades, report)

        wb = Workbook()

        # 1. 요약 시트
        self._create_summary_sheet(wb, report)

        # 2. 거래 내역 시트
        self._create_trades_sheet(wb, trades)

        # 3. 시간대별 분석 시트
        self._create_hourly_sheet(wb, report)

        # 4. 신호 분석 시트
        self._create_signal_sheet(wb, report)

        # 5. 청산 분석 시트
        self._create_exit_sheet(wb, report)

        # 기본 시트 제거
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # 저장
        if filename is None:
            filename = f"v7_backtest_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        output_path = self.config.output_dir / filename
        wb.save(output_path)
        self.logger.info(f"Excel 보고서 저장: {output_path}")

        return output_path

    def _create_summary_sheet(self, wb, report: Dict):
        """요약 시트 생성"""
        ws = wb.create_sheet("요약", 0)

        # 스타일
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)

        # 제목
        ws['A1'] = "V7 Purple 3분봉 백테스트 결과"
        ws['A1'].font = title_font
        ws['A2'] = f"기간: {self.config.event_start} ~ {self.config.event_end}"
        ws['A3'] = f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        # 기본 통계
        basic = report.get('basic_stats', {})

        row = 5
        ws.cell(row, 1, "기본 통계").font = header_font

        stats = [
            ("총 거래수", basic.get('total_trades', 0)),
            ("승리 거래", basic.get('winning_trades', 0)),
            ("패배 거래", basic.get('losing_trades', 0)),
            ("승률 (%)", basic.get('win_rate', 0)),
            ("", ""),
            ("총 수익률 (%)", basic.get('total_return_pct', 0)),
            ("평균 수익률 (%)", basic.get('avg_return_pct', 0)),
            ("평균 승리 (%)", basic.get('avg_win_pct', 0)),
            ("평균 패배 (%)", basic.get('avg_loss_pct', 0)),
            ("", ""),
            ("Profit Factor", basic.get('profit_factor', 0)),
            ("Expectancy (%)", basic.get('expectancy', 0)),
            ("최대 낙폭 (%)", basic.get('max_drawdown_pct', 0)),
            ("", ""),
            ("연속 최대 승리", basic.get('max_consecutive_wins', 0)),
            ("연속 최대 패배", basic.get('max_consecutive_losses', 0)),
            ("", ""),
            ("총 비용 전 손익", basic.get('total_gross_pnl', 0)),
            ("총 비용", basic.get('total_cost', 0)),
            ("순 손익", basic.get('total_net_pnl', 0)),
        ]

        for i, (label, value) in enumerate(stats):
            ws.cell(row + 1 + i, 1, label)
            ws.cell(row + 1 + i, 2, value)

        # 열 너비 조정
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15

    def _create_trades_sheet(self, wb, trades: List[Trade]):
        """거래 내역 시트 생성"""
        ws = wb.create_sheet("거래내역")

        # 헤더
        headers = [
            "종목코드", "종목명", "이벤트일", "진입시간", "진입가",
            "청산시간", "청산가", "청산유형", "수익률(%)", "R-Multiple",
            "MFE(%)", "MAE(%)", "보유봉수", "신호Score", "신호시간"
        ]

        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font

        # 데이터
        for row, trade in enumerate(trades, 2):
            ws.cell(row, 1, trade.stock_code)
            ws.cell(row, 2, trade.stock_name)
            ws.cell(row, 3, str(trade.event_date))
            ws.cell(row, 4, str(trade.entry_dt))
            ws.cell(row, 5, trade.entry_price)
            ws.cell(row, 6, str(trade.exit_dt))
            ws.cell(row, 7, trade.exit_price)
            ws.cell(row, 8, trade.exit_type)
            ws.cell(row, 9, trade.net_return_pct)
            ws.cell(row, 10, trade.r_multiple)
            ws.cell(row, 11, trade.mfe_pct)
            ws.cell(row, 12, trade.mae_pct)
            ws.cell(row, 13, trade.holding_bars)
            ws.cell(row, 14, trade.signal_score)
            ws.cell(row, 15, trade.signal_time)

            # 수익/손실 색상
            if trade.net_return_pct > 0:
                ws.cell(row, 9).font = Font(color="006400")  # 녹색
            else:
                ws.cell(row, 9).font = Font(color="8B0000")  # 적색

        # 열 너비
        col_widths = [10, 12, 12, 18, 10, 18, 10, 15, 10, 10, 8, 8, 8, 10, 8]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width

    def _create_hourly_sheet(self, wb, report: Dict):
        """시간대별 분석 시트 생성"""
        ws = wb.create_sheet("시간대별")

        hourly = report.get('hourly_analysis', {}).get('hourly', {})

        # 헤더
        headers = ["시간대", "거래수", "평균수익률(%)", "총수익률(%)", "승률(%)"]
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font

        # 데이터
        row = 2
        for time_slot, data in sorted(hourly.items()):
            ws.cell(row, 1, time_slot)
            ws.cell(row, 2, data.get('count', 0))
            ws.cell(row, 3, data.get('avg_return', 0))
            ws.cell(row, 4, data.get('total_return', 0))
            ws.cell(row, 5, data.get('win_rate', 0))
            row += 1

        # 최고/최저 시간대
        row += 1
        ws.cell(row, 1, "최고 성과 시간대").font = Font(bold=True)
        ws.cell(row, 2, f"{report.get('hourly_analysis', {}).get('best_hour', 0):02d}:00")

        row += 1
        ws.cell(row, 1, "최저 성과 시간대").font = Font(bold=True)
        ws.cell(row, 2, f"{report.get('hourly_analysis', {}).get('worst_hour', 0):02d}:00")

        # 열 너비
        for col in range(1, 6):
            ws.column_dimensions[chr(64 + col)].width = 15

    def _create_signal_sheet(self, wb, report: Dict):
        """신호 분석 시트 생성"""
        ws = wb.create_sheet("신호분석")

        signal = report.get('signal_analysis', {})

        # Score 통계
        ws['A1'] = "Score 통계"
        ws['A1'].font = Font(bold=True)

        score_stats = signal.get('score_stats', {})
        row = 2
        for key, value in score_stats.items():
            ws.cell(row, 1, key)
            ws.cell(row, 2, value)
            row += 1

        ws.cell(row, 1, "Score-수익률 상관계수")
        ws.cell(row, 2, signal.get('score_return_correlation', 0))

        # Score 구간별 성과
        row += 2
        ws.cell(row, 1, "Score 구간별 성과").font = Font(bold=True)
        row += 1

        headers = ["구간", "거래수", "평균수익률(%)", "승률(%)"]
        for col, header in enumerate(headers, 1):
            ws.cell(row, col, header).font = Font(bold=True)

        row += 1
        for bucket, data in signal.get('by_score', {}).items():
            ws.cell(row, 1, bucket)
            ws.cell(row, 2, data.get('count', 0))
            ws.cell(row, 3, data.get('avg_return', 0))
            ws.cell(row, 4, data.get('win_rate', 0))
            row += 1

        # Rise 구간별 성과
        row += 1
        ws.cell(row, 1, "Rise 구간별 성과").font = Font(bold=True)
        row += 1

        for col, header in enumerate(headers, 1):
            ws.cell(row, col, header).font = Font(bold=True)

        row += 1
        for bucket, data in signal.get('by_rise', {}).items():
            ws.cell(row, 1, bucket)
            ws.cell(row, 2, data.get('count', 0))
            ws.cell(row, 3, data.get('avg_return', 0))
            ws.cell(row, 4, data.get('win_rate', 0))
            row += 1

        # 열 너비
        for col in range(1, 5):
            ws.column_dimensions[chr(64 + col)].width = 15

    def _create_exit_sheet(self, wb, report: Dict):
        """청산 분석 시트 생성"""
        ws = wb.create_sheet("청산분석")

        exit_analysis = report.get('exit_analysis', {})

        # 헤더
        headers = ["청산유형", "거래수", "비율(%)", "평균수익률(%)", "승률(%)"]
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.fill = header_fill
            cell.font = header_font

        # 데이터
        row = 2
        for exit_type, data in exit_analysis.items():
            ws.cell(row, 1, exit_type)
            ws.cell(row, 2, data.get('count', 0))
            ws.cell(row, 3, data.get('pct', 0))
            ws.cell(row, 4, data.get('avg_return', 0))
            ws.cell(row, 5, data.get('win_rate', 0))
            row += 1

        # MFE/MAE 분석
        row += 2
        ws.cell(row, 1, "MFE/MAE 분석").font = Font(bold=True)

        mfe_mae = report.get('mfe_mae_analysis', {})

        row += 1
        ws.cell(row, 1, "전체").font = Font(bold=True)
        overall = mfe_mae.get('overall', {})
        row += 1
        ws.cell(row, 1, "평균 MFE (%)")
        ws.cell(row, 2, overall.get('avg_mfe', 0))
        row += 1
        ws.cell(row, 1, "평균 MAE (%)")
        ws.cell(row, 2, overall.get('avg_mae', 0))
        row += 1
        ws.cell(row, 1, "최대 MFE (%)")
        ws.cell(row, 2, overall.get('max_mfe', 0))
        row += 1
        ws.cell(row, 1, "최대 MAE (%)")
        ws.cell(row, 2, overall.get('max_mae', 0))

        row += 2
        ws.cell(row, 1, "승리 거래").font = Font(bold=True)
        winners = mfe_mae.get('winners', {})
        row += 1
        ws.cell(row, 1, "평균 MFE (%)")
        ws.cell(row, 2, winners.get('avg_mfe', 0))
        row += 1
        ws.cell(row, 1, "평균 MAE (%)")
        ws.cell(row, 2, winners.get('avg_mae', 0))

        row += 2
        ws.cell(row, 1, "패배 거래").font = Font(bold=True)
        losers = mfe_mae.get('losers', {})
        row += 1
        ws.cell(row, 1, "평균 MFE (%)")
        ws.cell(row, 2, losers.get('avg_mfe', 0))
        row += 1
        ws.cell(row, 1, "평균 MAE (%)")
        ws.cell(row, 2, losers.get('avg_mae', 0))

        # R-Multiple 분석
        row += 2
        ws.cell(row, 1, "R-Multiple 분석").font = Font(bold=True)
        r_analysis = report.get('r_multiple_analysis', {})
        r_stats = r_analysis.get('stats', {})

        row += 1
        ws.cell(row, 1, "평균")
        ws.cell(row, 2, r_stats.get('mean', 0))
        row += 1
        ws.cell(row, 1, "중앙값")
        ws.cell(row, 2, r_stats.get('median', 0))
        row += 1
        ws.cell(row, 1, "표준편차")
        ws.cell(row, 2, r_stats.get('std', 0))
        row += 1
        ws.cell(row, 1, "최소")
        ws.cell(row, 2, r_stats.get('min', 0))
        row += 1
        ws.cell(row, 1, "최대")
        ws.cell(row, 2, r_stats.get('max', 0))

        # R 분포
        row += 2
        ws.cell(row, 1, "R-Multiple 분포").font = Font(bold=True)
        r_dist = r_analysis.get('distribution', {})
        row += 1
        for bucket, count in r_dist.items():
            ws.cell(row, 1, bucket)
            ws.cell(row, 2, count)
            row += 1

        # 열 너비
        for col in range(1, 6):
            ws.column_dimensions[chr(64 + col)].width = 15

    def _export_csv_fallback(self, trades: List[Trade], report: Dict) -> Path:
        """openpyxl 미설치 시 CSV 대체"""
        # 거래 내역
        records = []
        for t in trades:
            records.append({
                "stock_code": t.stock_code,
                "stock_name": t.stock_name,
                "event_date": t.event_date,
                "entry_dt": t.entry_dt,
                "entry_price": t.entry_price,
                "exit_dt": t.exit_dt,
                "exit_price": t.exit_price,
                "exit_type": t.exit_type,
                "net_return_pct": t.net_return_pct,
                "r_multiple": t.r_multiple,
                "mfe_pct": t.mfe_pct,
                "mae_pct": t.mae_pct,
                "signal_score": t.signal_score
            })

        df = pd.DataFrame(records)
        output_path = self.config.trades_path
        df.to_csv(output_path, index=False, encoding="utf-8-sig")

        self.logger.info(f"CSV 대체 저장: {output_path}")
        return output_path
